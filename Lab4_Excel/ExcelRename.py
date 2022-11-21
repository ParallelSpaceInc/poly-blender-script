import os
import zipfile
import openpyxl

from genericpath import isfile
from openpyxl import Workbook
from openpyxl.styles import Alignment

file_loc_zip = "C:\\Users\\USER\\Downloads\\Remesh\\Renames\\"
file_loc_excel = 'C:\\Users\\USER\\Downloads\\Remesh\\Excel\\'

wb = openpyxl.load_workbook(file_loc_excel + "ReadTable.xlsx")

ws = wb.create_sheet('Models')
ws = wb.active

align = Alignment(horizontal='center', vertical='center')

size_model = []
name_model = []

file_list = os.listdir(file_loc_zip)
#print(file_list)

for model in file_list:
    name_model.append(os.path.splitext(model)[0].split('_', 1)[1])
    modelpath = file_loc_zip+model
    if os.path.isfile(modelpath):
        ziptmp = zipfile.ZipFile(modelpath)
        sizekb = ziptmp.getinfo(ziptmp.namelist()[1]).file_size // 1024
        size_model.append(sizekb)

for x in range(2, (len(file_list)+2)):
    ws.cell(x, 3, name_model[x-2])
    ws.cell(x, 9, size_model[x-2])

for row in ws.rows:
    for col in row[0:]:
        col.alignment = align

wb.save("C:/Users/USER/Downloads/Remesh/Excel/ModelTable3.xlsx")