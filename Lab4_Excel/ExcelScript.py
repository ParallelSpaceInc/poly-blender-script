import os
import zipfile

from genericpath import isfile
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()

ws = wb.create_sheet('Models')
ws = wb.active

align = Alignment(horizontal='center', vertical='center')

path_origin = "C:/Users/USER/Downloads/Remesh/Old/Origins/"
path_s100_r10_gltf = "C:/Users/USER/Downloads/Remesh/models_gltf_s100_r10_0902/"
path_s100_r10_glb = "C:/Users/USER/Downloads/Remesh/models_glb_s100_r10_0906/"
path_s500_r10_glb = "C:/Users/USER/Downloads/Remesh/models_glb_s500_r10_0919/"

size_origin = []
size_s100_r10_gltf = []
size_s100_r10_glb = []
size_s500_r10_glb = []

ws.append(["순번", "이름", "원본 PLY 용량(KB)", "s100_r10_gltf(KB)", "s100_r10_glb(KB)", "s500_r10_glb(KB)"])
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18

file_list = os.listdir(path_s500_r10_glb)

for model in file_list:
    modelpath = path_origin+os.path.splitext(model)[0]+'.ply'
    if os.path.isfile(modelpath):
        sizekb = os.path.getsize(modelpath) // 1024
        size_origin.append(sizekb)

    modelpath = path_s100_r10_gltf+model
    if os.path.isfile(modelpath):
        ziptmp = zipfile.ZipFile(modelpath)
        sizekb = ziptmp.getinfo(ziptmp.namelist()[1]).file_size // 1024
        size_s100_r10_gltf.append(sizekb)

    modelpath = path_s100_r10_glb+model
    if os.path.isfile(modelpath):
        ziptmp = zipfile.ZipFile(modelpath)
        sizekb = ziptmp.getinfo(ziptmp.namelist()[1]).file_size // 1024
        size_s100_r10_glb.append(sizekb)

    modelpath = path_s500_r10_glb+model
    if os.path.isfile(modelpath):
        ziptmp = zipfile.ZipFile(modelpath)
        sizekb = ziptmp.getinfo(ziptmp.namelist()[1]).file_size // 1024
        size_s500_r10_glb.append(sizekb)

for x in range(2, (len(file_list)+1)):
    ws.cell(x, 2, os.path.splitext(file_list[x-2])[0])
    ws.cell(x, 3, size_origin[x-2])
    ws.cell(x, 4, size_s100_r10_gltf[x-2])
    ws.cell(x, 5, size_s100_r10_glb[x-2])
    ws.cell(x, 6, size_s500_r10_glb[x-2])

for row in ws.rows:
    for col in row[0:]:
        col.alignment = align

wb.save("C:/Users/USER/Downloads/Remesh/Excel/ModelTable.xlsx")