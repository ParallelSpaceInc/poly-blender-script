import os
import bpy

from genericpath import isfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from bpy import context

wb = Workbook()

ws = wb.create_sheet('Models')
ws = wb.active

align = Alignment(horizontal='center', vertical='center')

path_origin = "C:/Users/USER/Downloads/Remesh/Old/Origins/"
path_s500_r10_glb = "C:/Users/USER/Downloads/Remesh/models_glb_s500_r10_0919/"

sizeX = []
sizeY = []
sizeZ = []

ws.append(["순번", "이름","sizeX", "sizeY", "sizeZ"])
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10

file_list = os.listdir(path_s500_r10_glb)

txt = open("C:/Users/USER/Downloads/Remesh/Excel/txt.txt", "w")

for model in file_list:
    modelpath = path_origin+os.path.splitext(model)[0]+'.ply'
    if os.path.isfile(modelpath):
        imported_object = bpy.ops.import_mesh.ply(filepath=modelpath)
        sizeX.append(bpy.context.object.dimensions.x)
        sizeY.append(bpy.context.object.dimensions.y)
        sizeZ.append(bpy.context.object.dimensions.z)
        txt.write(str(bpy.context.object.dimensions.x) + " " + str(bpy.context.object.dimensions.y) + " " + str(bpy.context.object.dimensions.z) + "\n")
        bpy.ops.object.delete()
        
txt.close()
for x in range(2, (len(file_list)+1)):
    ws.cell(x, 2, os.path.splitext(file_list[x-2])[0])
    ws.cell(x, 3, sizeX[x-2])
    ws.cell(x, 4, sizeY[x-2])
    ws.cell(x, 5, sizeZ[x-2])
    
for row in ws.rows:
    for col in row[0:]:
        col.alignment = align
        
wb.save("C:/Users/USER/Downloads/Remesh/Excel/ModelSize.xlsx")